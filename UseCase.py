import openpyxl, io
import Display_Visual_Charts_OPENPYXL as c0
import Charts_uisng_Panda_Matplotlib as c1

workbook1 = openpyxl.load_workbook("inventory.xlsx")
sheet1 = workbook1["Sheet1"]
product_per_supplier = {}
price_per_supplier = {}
product_l10_inventory = {}
f1 = open("c3_output.txt", "w")

# Exercise 1 - get count of products per supplier company
for product_row in range(2, sheet1.max_row + 1):
    supplier_name = sheet1.cell(product_row, 4).value
    if supplier_name in product_per_supplier:
        current_product_num = product_per_supplier.get(supplier_name)
        product_per_supplier[supplier_name] = current_product_num + 1
    else:
        product_per_supplier[supplier_name] = 1
print(f"Total product per supplier \n {product_per_supplier}", file=f1)

# Exercise 2 - get sum of price per company
for product_row_1 in range(2, sheet1.max_row + 1):
    supplier_name_1 = sheet1.cell(product_row_1, 4).value
    inventory = sheet1.cell(product_row_1, 2).value
    price = sheet1.cell(product_row_1, 3).value
    if supplier_name_1 in price_per_supplier:
        # print(price_per_supplier)
        price_per_supplier[supplier_name_1] = price_per_supplier[supplier_name_1] + inventory * price
    else:
        price_per_supplier[supplier_name_1] = inventory * price
print(f"Total Price per supplier \n {price_per_supplier}", file=f1)

# Exercise 3 - print all the products less than 10 inventories
for product_row_2 in range(2, sheet1.max_row + 1):
    product = int(sheet1.cell(product_row_2, 1).value)
    inventory = int(sheet1.cell(product_row_2, 2).value)
    if inventory < 10:
        product_l10_inventory[product] = inventory
print(f"Product inventories less than 10 \n {product_l10_inventory}", file=f1)

# Exercise 4 : Insert column 5 on spreadsheet as total price
header5 = sheet1.cell(1, 5, "Total_Price")
for product_row_3 in range(2, sheet1.max_row + 1):
    total_price = sheet1.cell(product_row_3, 5)
    inventory1 = int(sheet1.cell(product_row_3, 2).value)
    price1 = sheet1.cell(product_row_3, 3).value
    total_price.value = inventory1 * price1

workbook1.save("inventory_after_exercise.xlsx")  # this will edit the workbook and will add the 5th column
c0.chart_openpyxl()
c1.charts_function()
