import openpyxl
from openpyxl.chart import *


def chart_openpyxl():
    wb2 = openpyxl.load_workbook("inventory_after_exercise.xlsx")
    sheet1 = wb2["Sheet1"]

    cat = Reference(sheet1, min_col=4, max_col=4, min_row=2, max_row=sheet1.max_row)  # For x-axis - Supplier col 4
    values1 = Reference(sheet1, min_col=2, max_col=2, min_row=1, max_row=4)  # for Y-axis Inventory col-2
    # values2 = Reference(sheet2, min_col=3, max_col=3, min_row=1, max_row=sheet2.max_row)  # for Y-axis Inventory col-2
    chart = BarChart()
    chart1 = BarChart3D()
    chart2 = LineChart()

    chart.x_axis.title = "Supplier"
    chart.y_axis.title = "Inventory"
    chart.add_data(values1, titles_from_data=True)
    chart1.add_data(values1, titles_from_data=True)
    chart2.add_data(values1, titles_from_data=True)
    # series = Series(values=values2, xvalues=values1, title="Test", title_from_data=True)
    # chart.series.append(series)
    chart.set_categories(cat)
    chart1.set_categories(cat)
    chart2.set_categories(cat)
    chart.type = "col"  # col for horizontal and bar for vertical
    chart.grouping = "stacked"  # Value must be one of {‘standard’, ‘stacked’, ‘clustered’, ‘percentStacked’}
    chart.overlap = 100
    chart.title = " BAR-CHART "
    chart1.title = " BAR-CHART 3D "
    chart2.title = " LINE-CHART "
    sheet1.add_chart(chart, "G2")
    sheet1.add_chart(chart1, "G17")
    sheet1.add_chart(chart2, "G33")
    wb2.save("c0_openpyxl.xlsx")
